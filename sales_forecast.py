{\rtf1\ansi\ansicpg1251\cocoartf2821
\cocoatextscaling0\cocoaplatform0{\fonttbl\f0\fnil\fcharset0 Menlo-Regular;}
{\colortbl;\red255\green255\blue255;\red255\green255\blue255;\red0\green0\blue0;}
{\*\expandedcolortbl;;\cssrgb\c100000\c100000\c100000;\cssrgb\c0\c0\c0;}
\paperw11900\paperh16840\margl1440\margr1440\vieww11520\viewh8400\viewkind0
\deftab720
\pard\pardeftab720\partightenfactor0

\f0\fs26\fsmilli13333 \cf0 \cb2 \expnd0\expndtw0\kerning0
import openpyxl\
import numpy as np\
import pandas as pd\
from datetime import datetime\
from pmdarima import auto_arima\
from scipy.optimize import minimize\
import warnings\
\
warnings.filterwarnings("ignore")\
\
def load_sales_data(file_path):\
    """\uc0\u1047 \u1072 \u1075 \u1088 \u1091 \u1079 \u1082 \u1072  \u1076 \u1072 \u1085 \u1085 \u1099 \u1093  \u1080 \u1079  Excel (\u1087 \u1077 \u1088 \u1080 \u1086 \u1076 \u1099 , \u1094 \u1077 \u1085 \u1099 , \u1086 \u1073 \u1098 \u1077 \u1084 \u1099 , \u1080 \u1079 \u1076 \u1077 \u1088 \u1078 \u1082 \u1080 , \u1080 \u1085 \u1092 \u1083 \u1103 \u1094 \u1080 \u1103 )"""\
    try:\
        wb = openpyxl.load_workbook(file_path)\
        ws = wb.active\
        periods = [row.value for row in ws['A'][1:] if row.value]  # \uc0\u1055 \u1077 \u1088 \u1080 \u1086 \u1076 \u1099 \
        prices = [float(row.value) for row in ws['B'][1:] if row.value]  # \uc0\u1062 \u1077 \u1085 \u1099 \
        volumes = [float(row.value) for row in ws['C'][1:] if row.value]  # \uc0\u1054 \u1073 \u1098 \u1077 \u1084 \u1099 \
        costs = [float(row.value) for row in ws['D'][1:] if row.value]  # \uc0\u1048 \u1079 \u1076 \u1077 \u1088 \u1078 \u1082 \u1080 \
        # \uc0\u1048 \u1085 \u1092 \u1083 \u1103 \u1094 \u1080 \u1103  \u1080 \u1079  \u1103 \u1095 \u1077 \u1081 \u1082 \u1080  E1 (\u1075 \u1086 \u1076 \u1086 \u1074 \u1072 \u1103 , \u1074  \u1087 \u1088 \u1086 \u1094 \u1077 \u1085 \u1090 \u1072 \u1093 ), \u1077 \u1089 \u1083 \u1080  \u1085 \u1077 \u1090  \'97 \u1079 \u1072 \u1087 \u1088 \u1072 \u1096 \u1080 \u1074 \u1072 \u1077 \u1084  \u1091  \u1087 \u1086 \u1083 \u1100 \u1079 \u1086 \u1074 \u1072 \u1090 \u1077 \u1083 \u1103 \
        inflation_rate = ws['E1'].value if ws['E1'].value else float(input("\uc0\u1042 \u1074 \u1077 \u1076 \u1080 \u1090 \u1077  \u1075 \u1086 \u1076 \u1086 \u1074 \u1091 \u1102  \u1080 \u1085 \u1092 \u1083 \u1103 \u1094 \u1080 \u1102  (\u1074  %): ")) / 100\
        if len(periods) != len(prices) or len(periods) != len(volumes) or len(periods) != len(costs):\
            raise ValueError("\uc0\u1054 \u1096 \u1080 \u1073 \u1082 \u1072 : \u1076 \u1072 \u1085 \u1085 \u1099 \u1077  \u1085 \u1077 \u1082 \u1086 \u1088 \u1088 \u1077 \u1082 \u1090 \u1085 \u1099  \u1080 \u1083 \u1080  \u1086 \u1090 \u1089 \u1091 \u1090 \u1089 \u1090 \u1074 \u1091 \u1102 \u1090 .")\
        return periods, prices, volumes, costs, inflation_rate\
    except Exception as e:\
        print(f"\uc0\u1054 \u1096 \u1080 \u1073 \u1082 \u1072  \u1079 \u1072 \u1075 \u1088 \u1091 \u1079 \u1082 \u1080  \u1076 \u1072 \u1085 \u1085 \u1099 \u1093 : \{e\}")\
        return None, None, None, None, None\
\
def calculate_ema(series, period=3):\
    """\uc0\u1056 \u1072 \u1089 \u1095 \u1077 \u1090  EMA"""\
    return pd.Series(series).ewm(span=period, adjust=False).mean().iloc[-1]\
\
def adjust_for_inflation(value, months, inflation_rate):\
    """\uc0\u1050 \u1086 \u1088 \u1088 \u1077 \u1082 \u1090 \u1080 \u1088 \u1086 \u1074 \u1082 \u1072  \u1079 \u1085 \u1072 \u1095 \u1077 \u1085 \u1080 \u1103  \u1085 \u1072  \u1080 \u1085 \u1092 \u1083 \u1103 \u1094 \u1080 \u1102 """\
    monthly_inflation = (1 + inflation_rate) ** (1 / 12) - 1  # \uc0\u1052 \u1077 \u1089 \u1103 \u1095 \u1085 \u1072 \u1103  \u1080 \u1085 \u1092 \u1083 \u1103 \u1094 \u1080 \u1103 \
    return value * (1 + monthly_inflation) ** months\
\
def optimized_arima_forecast(series, periods, inflation_rate):\
    """\uc0\u1054 \u1087 \u1090 \u1080 \u1084 \u1080 \u1079 \u1080 \u1088 \u1086 \u1074 \u1072 \u1085 \u1085 \u1099 \u1081  ARIMA \u1089  \u1091 \u1095 \u1077 \u1090 \u1086 \u1084  \u1080 \u1085 \u1092 \u1083 \u1103 \u1094 \u1080 \u1080 """\
    try:\
        model = auto_arima(series, start_p=0, start_q=0, max_p=3, max_q=3, d=None,\
                          seasonal=True, m=12, stepwise=True, suppress_warnings=True,\
                          error_action='ignore', trace=False)\
        forecast = model.predict(n_periods=max(periods), return_conf_int=True)\
        forecasts = \{period: adjust_for_inflation(forecast[0][period-1], period, inflation_rate) \
                    for period in periods\}\
        conf_int = \{period: (adjust_for_inflation(forecast[1][period-1][0], period, inflation_rate),\
                            adjust_for_inflation(forecast[1][period-1][1], period, inflation_rate)) \
                    for period in periods\}\
        return forecasts, conf_int, model.order\
    except Exception as e:\
        print(f"\uc0\u1054 \u1096 \u1080 \u1073 \u1082 \u1072  ARIMA: \{e\}")\
        return None, None, (1, 1, 0)\
\
def calculate_volatility(series):\
    """\uc0\u1042 \u1086 \u1083 \u1072 \u1090 \u1080 \u1083 \u1100 \u1085 \u1086 \u1089 \u1090 \u1100  \u1076 \u1083 \u1103  \u1084 \u1077 \u1089 \u1103 \u1095 \u1085 \u1099 \u1093  \u1076 \u1072 \u1085 \u1085 \u1099 \u1093 """\
    log_returns = np.log(np.array(series[1:]) / np.array(series[:-1]))\
    return np.std(log_returns) * np.sqrt(12)\
\
def analyze_trend(series):\
    """\uc0\u1040 \u1085 \u1072 \u1083 \u1080 \u1079  \u1090 \u1088 \u1077 \u1085 \u1076 \u1072 """\
    x = np.arange(len(series))\
    slope = np.polyfit(x, series, 1)[0]\
    if slope > 0.01:\
        return "\uc0\u1042 \u1086 \u1089 \u1093 \u1086 \u1076 \u1103 \u1097 \u1080 \u1081  \u1090 \u1088 \u1077 \u1085 \u1076 "\
    elif slope < -0.01:\
        return "\uc0\u1053 \u1080 \u1089 \u1093 \u1086 \u1076 \u1103 \u1097 \u1080 \u1081  \u1090 \u1088 \u1077 \u1085 \u1076 "\
    else:\
        return "\uc0\u1057 \u1090 \u1072 \u1073 \u1080 \u1083 \u1100 \u1085 \u1099 \u1081  \u1090 \u1088 \u1077 \u1085 \u1076 "\
\
def calculate_correlation(prices, volumes):\
    """\uc0\u1050 \u1086 \u1088 \u1088 \u1077 \u1083 \u1103 \u1094 \u1080 \u1103  \u1084 \u1077 \u1078 \u1076 \u1091  \u1094 \u1077 \u1085 \u1086 \u1081  \u1080  \u1086 \u1073 \u1098 \u1077 \u1084 \u1086 \u1084 """\
    return np.corrcoef(prices, volumes)[0, 1]\
\
def estimate_demand_elasticity(prices, volumes):\
    """\uc0\u1054 \u1094 \u1077 \u1085 \u1082 \u1072  \u1101 \u1083 \u1072 \u1089 \u1090 \u1080 \u1095 \u1085 \u1086 \u1089 \u1090 \u1080  \u1089 \u1087 \u1088 \u1086 \u1089 \u1072 """\
    log_prices = np.log(prices)\
    log_volumes = np.log(volumes)\
    slope = np.polyfit(log_prices, log_volumes, 1)[0]\
    return slope\
\
def demand_function(price, elasticity, base_price, base_volume):\
    """\uc0\u1060 \u1091 \u1085 \u1082 \u1094 \u1080 \u1103  \u1089 \u1087 \u1088 \u1086 \u1089 \u1072 """\
    return base_volume * (price / base_price) ** elasticity\
\
def profit_function(price, elasticity, base_price, base_volume, cost_per_unit, months, inflation_rate):\
    """\uc0\u1060 \u1091 \u1085 \u1082 \u1094 \u1080 \u1103  \u1087 \u1088 \u1080 \u1073 \u1099 \u1083 \u1080  \u1089  \u1091 \u1095 \u1077 \u1090 \u1086 \u1084  \u1080 \u1085 \u1092 \u1083 \u1103 \u1094 \u1080 \u1080 """\
    volume = demand_function(price, elasticity, base_price, base_volume)\
    inflated_cost = adjust_for_inflation(cost_per_unit, months, inflation_rate)\
    return -(price - inflated_cost) * volume  # \uc0\u1054 \u1090 \u1088 \u1080 \u1094 \u1072 \u1090 \u1077 \u1083 \u1100 \u1085 \u1086 \u1077  \u1076 \u1083 \u1103  \u1084 \u1080 \u1085 \u1080 \u1084 \u1080 \u1079 \u1072 \u1094 \u1080 \u1080 \
\
def optimize_price_volume(prices, volumes, costs, price_forecasts, volume_forecasts, periods, inflation_rate):\
    """\uc0\u1044 \u1080 \u1085 \u1072 \u1084 \u1080 \u1095 \u1077 \u1089 \u1082 \u1072 \u1103  \u1086 \u1087 \u1090 \u1080 \u1084 \u1080 \u1079 \u1072 \u1094 \u1080 \u1103  \u1087 \u1088 \u1080 \u1073 \u1099 \u1083 \u1080  \u1089  \u1091 \u1095 \u1077 \u1090 \u1086 \u1084  \u1080 \u1085 \u1092 \u1083 \u1103 \u1094 \u1080 \u1080 """\
    elasticity = estimate_demand_elasticity(prices, volumes)\
    avg_cost = np.mean(costs)  # \uc0\u1057 \u1088 \u1077 \u1076 \u1085 \u1080 \u1077  \u1080 \u1079 \u1076 \u1077 \u1088 \u1078 \u1082 \u1080  \u1082 \u1072 \u1082  \u1073 \u1072 \u1079 \u1086 \u1074 \u1099 \u1077 \
    optimal_results = \{\}\
    \
    for period in periods:\
        base_price = price_forecasts[period]\
        base_volume = volume_forecasts[period]\
        bounds = [(base_price * 0.8, base_price * 1.2)]\
        \
        result = minimize(profit_function, x0=[base_price], \
                         args=(elasticity, base_price, base_volume, avg_cost, period, inflation_rate),\
                         bounds=bounds, method='L-BFGS-B')\
        \
        optimal_price = result.x[0]\
        optimal_volume = demand_function(optimal_price, elasticity, base_price, base_volume)\
        inflated_cost = adjust_for_inflation(avg_cost, period, inflation_rate)\
        optimal_profit = (optimal_price - inflated_cost) * optimal_volume\
        optimal_results[period] = (optimal_price, optimal_volume, optimal_profit)\
    \
    return optimal_results, elasticity, avg_cost\
\
def process_sales_forecast(file_path):\
    """\uc0\u1054 \u1073 \u1088 \u1072 \u1073 \u1086 \u1090 \u1082 \u1072  \u1076 \u1072 \u1085 \u1085 \u1099 \u1093  \u1089  \u1091 \u1095 \u1077 \u1090 \u1086 \u1084  \u1080 \u1085 \u1092 \u1083 \u1103 \u1094 \u1080 \u1080 """\
    try:\
        # \uc0\u1047 \u1072 \u1075 \u1088 \u1091 \u1079 \u1082 \u1072  \u1076 \u1072 \u1085 \u1085 \u1099 \u1093 \
        periods, prices, volumes, costs, inflation_rate = load_sales_data(file_path)\
        if not prices:\
            return\
        \
        # \uc0\u1040 \u1085 \u1072 \u1083 \u1080 \u1079  \u1076 \u1072 \u1085 \u1085 \u1099 \u1093 \
        price_trend = analyze_trend(prices)\
        volume_trend = analyze_trend(volumes)\
        price_volatility = calculate_volatility(prices)\
        volume_volatility = calculate_volatility(volumes)\
        price_ema = calculate_ema(prices)\
        volume_ema = calculate_ema(volumes)\
        correlation = calculate_correlation(prices, volumes)\
        avg_cost = np.mean(costs)\
        \
        # \uc0\u1055 \u1088 \u1086 \u1075 \u1085 \u1086 \u1079  \u1089  \u1091 \u1095 \u1077 \u1090 \u1086 \u1084  \u1080 \u1085 \u1092 \u1083 \u1103 \u1094 \u1080 \u1080 \
        periods_forecast = [1, 3, 6, 12]\
        price_forecasts, price_conf_int, price_arima_order = optimized_arima_forecast(prices, periods_forecast, inflation_rate)\
        volume_forecasts, volume_conf_int, volume_arima_order = optimized_arima_forecast(volumes, periods_forecast, inflation_rate)\
        \
        if price_forecasts is None:\
            price_forecasts = \{p: adjust_for_inflation(prices[-1], p, inflation_rate) for p in periods_forecast\}\
            price_conf_int = \{p: (prices[-1]*0.95, prices[-1]*1.05) for p in periods_forecast\}\
        if volume_forecasts is None:\
            volume_forecasts = \{p: volumes[-1] for p in periods_forecast\}\
            volume_conf_int = \{p: (volumes[-1]*0.95, volumes[-1]*1.05) for p in periods_forecast\}\
        \
        # \uc0\u1054 \u1087 \u1090 \u1080 \u1084 \u1080 \u1079 \u1072 \u1094 \u1080 \u1103 \
        optimal_results, elasticity, avg_cost = optimize_price_volume(\
            prices, volumes, costs, price_forecasts, volume_forecasts, periods_forecast, inflation_rate)\
        \
        # \uc0\u1055 \u1086 \u1076 \u1075 \u1086 \u1090 \u1086 \u1074 \u1082 \u1072  Excel\
        wb = openpyxl.load_workbook(file_path)\
        ws = wb.active\
        ws.title = "\uc0\u1055 \u1088 \u1086 \u1075 \u1085 \u1086 \u1079  \u1080  \u1086 \u1087 \u1090 \u1080 \u1084 \u1080 \u1079 \u1072 \u1094 \u1080 \u1103 "\
        \
        # \uc0\u1047 \u1072 \u1087 \u1080 \u1089 \u1100  \u1076 \u1072 \u1085 \u1085 \u1099 \u1093 \
        ws['E1'] = "\uc0\u1055 \u1077 \u1088 \u1080 \u1086 \u1076  (\u1084 \u1077 \u1089 \u1103 \u1094 \u1099 )"\
        ws['F1'] = "\uc0\u1055 \u1088 \u1086 \u1075 \u1085 \u1086 \u1079  \u1094 \u1077 \u1085 \u1099  (\u1089  \u1080 \u1085 \u1092 \u1083 .)"\
        ws['G1'] = "\uc0\u1044 \u1086 \u1074 . \u1080 \u1085 \u1090 \u1077 \u1088 \u1074 \u1072 \u1083  \u1094 \u1077 \u1085 \u1099 "\
        ws['H1'] = "\uc0\u1055 \u1088 \u1086 \u1075 \u1085 \u1086 \u1079  \u1086 \u1073 \u1098 \u1077 \u1084 \u1072 "\
        ws['I1'] = "\uc0\u1044 \u1086 \u1074 . \u1080 \u1085 \u1090 \u1077 \u1088 \u1074 \u1072 \u1083  \u1086 \u1073 \u1098 \u1077 \u1084 \u1072 "\
        ws['J1'] = "\uc0\u1054 \u1087 \u1090 \u1080 \u1084 \u1072 \u1083 \u1100 \u1085 \u1072 \u1103  \u1094 \u1077 \u1085 \u1072 "\
        ws['K1'] = "\uc0\u1054 \u1087 \u1090 \u1080 \u1084 \u1072 \u1083 \u1100 \u1085 \u1099 \u1081  \u1086 \u1073 \u1098 \u1077 \u1084 "\
        ws['L1'] = "\uc0\u1052 \u1072 \u1082 \u1089 . \u1087 \u1088 \u1080 \u1073 \u1099 \u1083 \u1100 "\
        ws['M1'] = "\uc0\u1040 \u1085 \u1072 \u1083 \u1080 \u1090 \u1080 \u1082 \u1072  \u1080  \u1088 \u1077 \u1082 \u1086 \u1084 \u1077 \u1085 \u1076 \u1072 \u1094 \u1080 \u1080 "\
        \
        row = 2\
        for period in periods_forecast:\
            period_name = format_period_name(period)\
            ws[f'E\{row\}'] = period_name\
            ws[f'F\{row\}'] = round(price_forecasts[period], 2)\
            ws[f'G\{row\}'] = f"[\{round(price_conf_int[period][0], 2)\}; \{round(price_conf_int[period][1], 2)\}]"\
            ws[f'H\{row\}'] = round(volume_forecasts[period], 2)\
            ws[f'I\{row\}'] = f"[\{round(volume_conf_int[period][0], 2)\}; \{round(volume_conf_int[period][1], 2)\}]"\
            ws[f'J\{row\}'] = round(optimal_results[period][0], 2)\
            ws[f'K\{row\}'] = round(optimal_results[period][1], 2)\
            ws[f'L\{row\}'] = round(optimal_results[period][2], 2)\
            row += 1\
        \
        # \uc0\u1040 \u1085 \u1072 \u1083 \u1080 \u1090 \u1080 \u1082 \u1072  \u1080  \u1088 \u1077 \u1082 \u1086 \u1084 \u1077 \u1085 \u1076 \u1072 \u1094 \u1080 \u1080 \
        ws['M2'] = f"\uc0\u1052 \u1086 \u1076 \u1077 \u1083 \u1100  \u1094 \u1077 \u1085 : ARIMA\{price_arima_order\}"\
        ws['M3'] = f"\uc0\u1052 \u1086 \u1076 \u1077 \u1083 \u1100  \u1086 \u1073 \u1098 \u1077 \u1084 \u1086 \u1074 : ARIMA\{volume_arima_order\}"\
        ws['M4'] = f"\uc0\u1058 \u1088 \u1077 \u1085 \u1076  \u1094 \u1077 \u1085 : \{price_trend\}"\
        ws['M5'] = f"\uc0\u1058 \u1088 \u1077 \u1085 \u1076  \u1086 \u1073 \u1098 \u1077 \u1084 \u1086 \u1074 : \{volume_trend\}"\
        ws['M6'] = f"EMA \uc0\u1094 \u1077 \u1085  (3): \{price_ema:.2f\}"\
        ws['M7'] = f"EMA \uc0\u1086 \u1073 \u1098 \u1077 \u1084 \u1086 \u1074  (3): \{volume_ema:.2f\}"\
        ws['M8'] = f"\uc0\u1042 \u1086 \u1083 \u1072 \u1090 \u1080 \u1083 \u1100 \u1085 \u1086 \u1089 \u1090 \u1100  \u1094 \u1077 \u1085 : \{price_volatility:.2%\}"\
        ws['M9'] = f"\uc0\u1042 \u1086 \u1083 \u1072 \u1090 \u1080 \u1083 \u1100 \u1085 \u1086 \u1089 \u1090 \u1100  \u1086 \u1073 \u1098 \u1077 \u1084 \u1086 \u1074 : \{volume_volatility:.2%\}"\
        ws['M10'] = f"\uc0\u1050 \u1086 \u1088 \u1088 \u1077 \u1083 \u1103 \u1094 \u1080 \u1103  \u1094 \u1077 \u1085 \u1072 /\u1086 \u1073 \u1098 \u1077 \u1084 : \{correlation:.2f\}"\
        ws['M11'] = f"\uc0\u1069 \u1083 \u1072 \u1089 \u1090 \u1080 \u1095 \u1085 \u1086 \u1089 \u1090 \u1100  \u1089 \u1087 \u1088 \u1086 \u1089 \u1072 : \{elasticity:.2f\}"\
        ws['M12'] = f"\uc0\u1057 \u1088 \u1077 \u1076 \u1085 \u1080 \u1077  \u1080 \u1079 \u1076 \u1077 \u1088 \u1078 \u1082 \u1080 : \{avg_cost:.2f\}"\
        ws['M13'] = f"\uc0\u1043 \u1086 \u1076 \u1086 \u1074 \u1072 \u1103  \u1080 \u1085 \u1092 \u1083 \u1103 \u1094 \u1080 \u1103 : \{inflation_rate*100:.2f\}%"\
        ws['M14'] = f"\uc0\u1044 \u1072 \u1090 \u1072  \u1088 \u1072 \u1089 \u1095 \u1077 \u1090 \u1072 : \{datetime.now().strftime('%Y-%m-%d')\}"\
        ws['M15'] = "\uc0\u1056 \u1077 \u1082 \u1086 \u1084 \u1077 \u1085 \u1076 \u1072 \u1094 \u1080 \u1080 :"\
        ws['M16'] = ("\uc0\u1057 \u1085 \u1080 \u1079 \u1080 \u1090 \u1100  \u1094 \u1077 \u1085 \u1091  \u1076 \u1083 \u1103  \u1088 \u1086 \u1089 \u1090 \u1072  \u1086 \u1073 \u1098 \u1077 \u1084 \u1072 " if elasticity < -1 else \
                    "\uc0\u1055 \u1086 \u1076 \u1085 \u1103 \u1090 \u1100  \u1094 \u1077 \u1085 \u1091  \u1089  \u1091 \u1095 \u1077 \u1090 \u1086 \u1084  \u1080 \u1085 \u1092 \u1083 \u1103 \u1094 \u1080 \u1080 " if correlation > -0.5 else \
                    "\uc0\u1057 \u1090 \u1072 \u1073 \u1080 \u1083 \u1080 \u1079 \u1080 \u1088 \u1086 \u1074 \u1072 \u1090 \u1100  \u1087 \u1088 \u1077 \u1076 \u1083 \u1086 \u1078 \u1077 \u1085 \u1080 \u1077 ")\
\
        # \uc0\u1057 \u1086 \u1093 \u1088 \u1072 \u1085 \u1077 \u1085 \u1080 \u1077 \
        wb.save(file_path)\
        print(f"\uc0\u1055 \u1088 \u1086 \u1075 \u1085 \u1086 \u1079  \u1080  \u1086 \u1087 \u1090 \u1080 \u1084 \u1080 \u1079 \u1072 \u1094 \u1080 \u1103  \u1091 \u1089 \u1087 \u1077 \u1096 \u1085 \u1086  \u1079 \u1072 \u1087 \u1080 \u1089 \u1072 \u1085 \u1099  \u1074  \u1092 \u1072 \u1081 \u1083 : \{file_path\}")\
        print(f"\uc0\u1058 \u1088 \u1077 \u1085 \u1076  \u1094 \u1077 \u1085 : \{price_trend\}")\
        print(f"\uc0\u1058 \u1088 \u1077 \u1085 \u1076  \u1086 \u1073 \u1098 \u1077 \u1084 \u1086 \u1074 : \{volume_trend\}")\
        print(f"EMA \uc0\u1094 \u1077 \u1085  (3): \{price_ema:.2f\}")\
        print(f"EMA \uc0\u1086 \u1073 \u1098 \u1077 \u1084 \u1086 \u1074  (3): \{volume_ema:.2f\}")\
        print(f"\uc0\u1042 \u1086 \u1083 \u1072 \u1090 \u1080 \u1083 \u1100 \u1085 \u1086 \u1089 \u1090 \u1100  \u1094 \u1077 \u1085 : \{price_volatility:.2%\}")\
        print(f"\uc0\u1042 \u1086 \u1083 \u1072 \u1090 \u1080 \u1083 \u1100 \u1085 \u1086 \u1089 \u1090 \u1100  \u1086 \u1073 \u1098 \u1077 \u1084 \u1086 \u1074 : \{volume_volatility:.2%\}")\
        print(f"\uc0\u1050 \u1086 \u1088 \u1088 \u1077 \u1083 \u1103 \u1094 \u1080 \u1103  \u1094 \u1077 \u1085 \u1072 /\u1086 \u1073 \u1098 \u1077 \u1084 : \{correlation:.2f\}")\
        print(f"\uc0\u1069 \u1083 \u1072 \u1089 \u1090 \u1080 \u1095 \u1085 \u1086 \u1089 \u1090 \u1100  \u1089 \u1087 \u1088 \u1086 \u1089 \u1072 : \{elasticity:.2f\}")\
        print(f"\uc0\u1057 \u1088 \u1077 \u1076 \u1085 \u1080 \u1077  \u1080 \u1079 \u1076 \u1077 \u1088 \u1078 \u1082 \u1080 : \{avg_cost:.2f\}")\
        print(f"\uc0\u1043 \u1086 \u1076 \u1086 \u1074 \u1072 \u1103  \u1080 \u1085 \u1092 \u1083 \u1103 \u1094 \u1080 \u1103 : \{inflation_rate*100:.2f\}%")\
        print("\\n\uc0\u1055 \u1088 \u1086 \u1075 \u1085 \u1086 \u1079  \u1080  \u1086 \u1087 \u1090 \u1080 \u1084 \u1072 \u1083 \u1100 \u1085 \u1099 \u1077  \u1079 \u1085 \u1072 \u1095 \u1077 \u1085 \u1080 \u1103  \u1087 \u1086  \u1087 \u1077 \u1088 \u1080 \u1086 \u1076 \u1072 \u1084  (\u1089  \u1091 \u1095 \u1077 \u1090 \u1086 \u1084  \u1080 \u1085 \u1092 \u1083 \u1103 \u1094 \u1080 \u1080 ):")\
        for period in periods_forecast:\
            print(f"\{format_period_name(period)\}: \uc0\u1055 \u1088 \u1086 \u1075 \u1085 \u1086 \u1079  \u1094 \u1077 \u1085 \u1099  = \{price_forecasts[period]:.2f\} "\
                  f"[\{price_conf_int[period][0]:.2f\}; \{price_conf_int[period][1]:.2f\}], "\
                  f"\uc0\u1055 \u1088 \u1086 \u1075 \u1085 \u1086 \u1079  \u1086 \u1073 \u1098 \u1077 \u1084 \u1072  = \{volume_forecasts[period]:.2f\} "\
                  f"[\{volume_conf_int[period][0]:.2f\}; \{volume_conf_int[period][1]:.2f\}], "\
                  f"\uc0\u1054 \u1087 \u1090 . \u1094 \u1077 \u1085 \u1072  = \{optimal_results[period][0]:.2f\}, "\
                  f"\uc0\u1054 \u1087 \u1090 . \u1086 \u1073 \u1098 \u1077 \u1084  = \{optimal_results[period][1]:.2f\}, "\
                  f"\uc0\u1055 \u1088 \u1080 \u1073 \u1099 \u1083 \u1100  = \{optimal_results[period][2]:.2f\}")\
        \
    except Exception as e:\
        print(f"\uc0\u1054 \u1096 \u1080 \u1073 \u1082 \u1072 : \{e\}")\
\
def format_period_name(months):\
    """\uc0\u1060 \u1086 \u1088 \u1084 \u1072 \u1090 \u1080 \u1088 \u1086 \u1074 \u1072 \u1085 \u1080 \u1077  \u1087 \u1077 \u1088 \u1080 \u1086 \u1076 \u1072 """\
    if months == 1:\
        return "1 \uc0\u1084 \u1077 \u1089 \u1103 \u1094 "\
    elif months == 3:\
        return "3 \uc0\u1084 \u1077 \u1089 \u1103 \u1094 \u1072 "\
    elif months == 6:\
        return "6 \uc0\u1084 \u1077 \u1089 \u1103 \u1094 \u1077 \u1074 "\
    elif months == 12:\
        return "1 \uc0\u1075 \u1086 \u1076 "\
    return f"\{months\} \uc0\u1084 \u1077 \u1089 \u1103 \u1094 \u1077 \u1074 "\
\
def main():\
    """\uc0\u1054 \u1089 \u1085 \u1086 \u1074 \u1085 \u1072 \u1103  \u1092 \u1091 \u1085 \u1082 \u1094 \u1080 \u1103 """\
    print("\uc0\u1055 \u1088 \u1086 \u1075 \u1088 \u1072 \u1084 \u1084 \u1072  \u1087 \u1088 \u1086 \u1075 \u1085 \u1086 \u1079 \u1072  \u1080  \u1086 \u1087 \u1090 \u1080 \u1084 \u1080 \u1079 \u1072 \u1094 \u1080 \u1080  \u1087 \u1088 \u1086 \u1076 \u1072 \u1078  (\u1089  \u1091 \u1095 \u1077 \u1090 \u1086 \u1084  \u1080 \u1085 \u1092 \u1083 \u1103 \u1094 \u1080 \u1080 )")\
    file_path = input("\uc0\u1042 \u1074 \u1077 \u1076 \u1080 \u1090 \u1077  \u1087 \u1091 \u1090 \u1100  \u1082  Excel-\u1092 \u1072 \u1081 \u1083 \u1091  \u1089  \u1076 \u1072 \u1085 \u1085 \u1099 \u1084 \u1080  (\u1085 \u1072 \u1087 \u1088 \u1080 \u1084 \u1077 \u1088 , 'sales.xlsx'): ")\
    process_sales_forecast(file_path)\
\
if __name__ == "__main__":\
    main()}